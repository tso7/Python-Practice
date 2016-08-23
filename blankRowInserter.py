#! /usr/bin/python3
# blankRowInserter.py - Insert M blank rows at row N in specified spreadsheet

import openpyxl, sys

# Retrieve command line arguments
if len(sys.argv) != 4:
    print('Invalid arguments. Usage: blankRowInserter.py <N> <M> <filename>')
    sys.exit()
dummy, N, M, filename = sys.argv

# Open workbook
wb = openpyxl.load_workbook('./excel/' + filename)
sheet = wb.active
maxRow = sheet.max_row
maxCol = sheet.max_column
sheetNew = wb.create_sheet(title=sheet.title+'new')

# Copy header and first N rows
print('In progress...')
for rowNum in range(1, int(N) + 2):
    for colNum in range(1, maxCol + 1):
        sheetNew.cell(row=rowNum, column = colNum).value =\
        sheet.cell(row=rowNum, column=colNum).value

# Insert remaining rows after inserting offset of M
for rowNum in range(int(N)+ 2, maxRow + 1):
    for colNum in range(1, maxCol + 1):
        sheetNew.cell(row=rowNum+int(M), column = colNum).value =\
        sheet.cell(row=rowNum, column=colNum).value

temp = sheet.title
wb.remove_sheet(sheet)
sheetNew.title = temp
wb.save('./excel/blankInserted.xlsx')
print('Done.')
