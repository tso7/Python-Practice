#! /usr/bin/python3
# transposeCells.py - Perform transpose of cells manually

import openpyxl, sys

# Retrieve filename
if len(sys.argv) != 2:
    print('Invalid arguments. Usage: transposeCells.py <filename>')
    sys.exit()
filename = sys.argv[1]

# Open workbook
wb = openpyxl.load_workbook('./excel/' + filename)
sheet = wb.active
maxRow = sheet.max_row
maxCol = sheet.max_column
sheetNew = wb.create_sheet(title=sheet.title+'_transposed')

# Copy header and first N rows
print('In progress...')
for rowNum in range(1, maxRow + 1):
    for colNum in range(1, maxCol + 1):
        sheetNew.cell(row=colNum, column = rowNum).value =\
        sheet.cell(row=rowNum, column=colNum).value

wb.save('./excel/' + filename)
print('Done.')
