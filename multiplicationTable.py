#! /usr/bin/python3
# multiplicationTable.py - Takes N from command line and generates excel sheet\
# with multiplication table of N

import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.cell import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active

# Retrieve value of N
if len(sys.argv) != 2:
    print('Invalid argument. Usage: multiplicationTable.py <N>')
    sys.exit()
N = sys.argv[1]

# Generate base values for first row and first column
boldFont = Font(size=12, bold=True)
for i in range(1, int(N) + 1):
    sheet.cell(row=i+1, column=1).font = boldFont
    sheet.cell(row=i+1, column=1).value = i
    sheet.cell(row=1, column=i+1).font = boldFont
    sheet.cell(row=1, column=i+1).value = i

# Generate multplication values based on row and column combinations
for i in range(2, int(N) + 2):
    for j in range (2, int(N) + 2):
        sheet.cell(row=i, column=j).value = '=' + get_column_letter(i) + '1*'\
        + get_column_letter(j) + '1'

wb.save('./excel/multiplicationTable.xlsx')
